import decimal
from django.http import HttpResponseForbidden
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.views.generic import CreateView
from django.contrib.auth.forms import UserCreationForm
from .models import *
from products.models import *
from django.http import HttpResponse
from openpyxl import Workbook
from Django_aboba.settings import EMAIL_HOST_USER
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from django.contrib.auth import authenticate, login, logout
import random
import string
import openpyxl
from notifications.models import Notification

@login_required
def home(request):
    worker, created = Worker.objects.get_or_create(user=request.user)
    first_name = worker.first_name
    last_name = worker.last_name
    email_ = worker.email
    money = worker.money
    speciality = worker.speciality
    context = {"first_name": first_name,
               "last_name": last_name,
               "email": email_,
               "money": money,
               "spec": speciality
               }
    return render(request, "main/home.html", context=context)


@login_required
def shop_edit(request):
    worker, _ = Worker.objects.get_or_create(user=request.user)
    first_name = worker.first_name
    last_name = worker.last_name
    surname = worker.surname
    email_ = worker.email
    money = worker.money
    speciality = worker.speciality
    birthday = worker.birthday

    products = Product.objects.filter(is_active=True).order_by('price')
    is_admin = request.user.is_staff
    purchases = request.user.purchases.all().order_by('-date')
    wishlist, _ = Wishlist.objects.get_or_create(user=request.user)

    context = {"admin": is_admin,
               "first_name": first_name,
               "last_name": last_name,
               "email": email_,
               "money": money,
               "birthday": birthday,
               "surname": surname,
               "spec": speciality,
               "products": products,
               "purchases": purchases,
               "wishlist": wishlist.products.all(),
               }
    return render(request, 'main/home_edit.html', context=context)

@login_required
def user(request):
    worker = request.user.worker
    products = Product.objects.all()
    purchases = request.user.purchases.all()
    wishlist, _ = Wishlist.objects.get_or_create(user=request.user)

    context = {
        'user': worker,
        'products': products,
        'purchases': purchases,
        'wishlist': wishlist.products.all()
    }
    return render(request, 'landing/user.html', context)


@login_required
def product_purchase(request, pk):
    product = get_object_or_404(Product, pk=pk)
    worker = request.user.worker
    wishlist, _ = Wishlist.objects.get_or_create(user=request.user)
    if product.price <= worker.money:
        worker.money -= product.price
        letters = string.ascii_lowercase
        purchase = Purchase(product=product, user=request.user, promocod=''.join(random.choice(letters) for i in range(10)))
        if product in wishlist.products.all():
            wishlist.products.remove(product)
            wishlist.save()
        purchase.save()
        worker.save()
    return redirect('/cart')


@login_required
def add_to_wishlist(request, pk):
    product = get_object_or_404(Product, pk=pk)
    wishlist, _ = Wishlist.objects.get_or_create(user=request.user)
    if product not in wishlist.products.all():
        wishlist.products.add(product)
    return redirect('/home')


@login_required
def remove_from_wishlist(request, pk):
    product = get_object_or_404(Product, pk=pk)
    wishlist = get_object_or_404(Wishlist, user=request.user)
    wishlist.products.remove(product)
    return redirect('/cart')


class SignUp(CreateView):
    form_class = UserCreationForm
    success_url = reverse_lazy("login")
    template_name = "registration/login.html"

@login_required
def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def shop(request):
    worker, _ = Worker.objects.get_or_create(user=request.user)
    first_name = worker.first_name
    last_name = worker.last_name
    surname = worker.surname
    email_ = worker.email
    money = worker.money
    speciality = worker.speciality
    birthday = worker.birthday

    products = Product.objects.filter(is_active=True).order_by('price')
    is_admin = request.user.is_staff
    purchases = request.user.purchases.all().order_by('-date')
    wishlist, _ = Wishlist.objects.get_or_create(user=request.user)

    context = {"admin": is_admin,
               "first_name": first_name,
               "last_name": last_name,
               "email": email_,
               "money": money,
               "birthday": birthday,
               "surname": surname,
               "spec": speciality,
               "products": products,
               "purchases": purchases,
               "wishlist": wishlist.products.all(),
               }
    return render(request, 'main/home.html', context=context)


@login_required
def cart(request):
    worker, _ = Worker.objects.get_or_create(user=request.user)
    first_name = worker.first_name
    last_name = worker.last_name
    surname = worker.surname
    email_ = worker.email
    money = worker.money
    speciality = worker.speciality
    birthday = worker.birthday

    products = Product.objects.filter(is_active=True).order_by('price')
    is_admin = request.user.is_staff
    purchases = request.user.purchases.all().order_by('-date')
    wishlist, _ = Wishlist.objects.get_or_create(user=request.user)

    context = {"admin": is_admin,
               "first_name": first_name,
               "last_name": last_name,
               "email": email_,
               "money": money,
               "birthday": birthday,
               "surname": surname,
               "spec": speciality,
               "products": products,
               "purchases": purchases,
               "wishlist": wishlist.products.all(),
               }
    return render(request, 'main/cart.html', context=context)


@login_required
def staff(request):
    if not request.user.is_staff:
        return redirect('/home')

    worker, _ = Worker.objects.get_or_create(user=request.user)
    first_name = worker.first_name
    last_name = worker.last_name
    surname = worker.surname
    email_ = worker.email
    money = worker.money
    speciality = worker.speciality
    birthday = worker.birthday
    workers = Worker.objects.all()
    notifications = Notification.objects.filter(notification_receiver=worker, was_seen=False)
    seen_notifications = Notification.objects.filter(notification_receiver=worker, was_seen=True)
    notifications_count = notifications.count()

    context = {"first_name": first_name,
               "last_name": last_name,
               "email": email_,
               "money": money,
               "birthday": birthday,
               "surname": surname,
               "spec": speciality,
               "workers": workers,
               "notifications_count": notifications_count,
               "notifications": notifications,
               "seen_notifications": seen_notifications
               }
    return render(request, 'main/staff.html', context=context)


@login_required
def receipt_add(request):
    if request.method == 'POST':
        worker = Worker.objects.get_or_create(user=request.user)[0]
        receipt = Receipt(image=request.FILES['file'])
        receipt.worker = worker
        receipt.save()
        staff_users = User.objects.filter(is_staff=True)
        admin_receivers = []
        for work in staff_users:
            if worker.city == work.worker.city:
                admin_receivers.append(work)
        if len(admin_receivers) == 0:
            for work in staff_users:
                admin_receivers.append(work)
        for obj in admin_receivers:
            new_notification = Notification.objects.create(received_receipt=receipt, notification_receiver=obj.worker)
            new_notification.save()
        return redirect('/cart')


@login_required
def remove_notification(request, pk):
    note =  get_object_or_404(Notification, pk=pk)
    note.was_seen = True
    return redirect('/staff')


@login_required
def create_worker(request):
    if request.method == 'POST':
        last_name = request.POST.get('last_name')
        first_name = request.POST.get('first_name')
        surname = request.POST.get('patronymic')
        speciality = request.POST.get('speciality')
        birthday = request.POST.get('birthday')
        username = request.POST.get('username')
        password = request.POST.get('password')
        city = request.POST.get('city')
        email = request.POST.get('email')

        if username == "" or password == "":
            return redirect('/staff')

        if User.objects.filter(username=username).exists():
            return redirect('/staff')

        new_user = User.objects.create_user(
            username=username,
            password=password,
        )
        new_user.save()

        worker, _ = Worker.objects.get_or_create(user=new_user)
        worker.first_name = first_name
        worker.last_name = last_name
        worker.surname = surname
        worker.speciality = speciality
        worker.birthday = birthday
        worker.email = email
        worker.city = city
        worker.save()

    return redirect('/staff')


@login_required
def delete_worker(request):
    if request.method == 'POST':
        username = request.POST.get('username')

        if User.objects.filter(username=username).exists():
            user = User.objects.get(username=username)
            user.delete()

    return redirect('/staff')


@login_required
def edit_worker(request):
    if request.method == 'POST':
        new_firstname = request.POST.get('new_firstname')
        new_lastname = request.POST.get('new_lastname')
        new_surname = request.POST.get('new_surname')
        new_speciality = request.POST.get('new_speciality')
        new_birthday = request.POST.get('new_birthday')
        new_email = request.POST.get('new_email')
        new_city = request.POST.get('new_city')
        money_added = request.POST.get('money')
        pk = request.POST.get('worker_id')
        worker = get_object_or_404(Worker, pk=pk)

        if new_firstname != "":
            worker.first_name = new_firstname
        if new_lastname != "":
            worker.last_name = new_lastname
        if new_surname != "":
            worker.surname = new_surname
        if new_speciality != "":
            worker.speciality = new_speciality
        if new_birthday != "":
            worker.birthday = new_birthday
        if new_email != "":
            worker.email = new_email
        if new_city != "":
            worker.city = new_city
        if money_added != "":
            worker.money += decimal.Decimal(money_added)
        worker.save()

    return redirect('/staff')


@login_required
def create_product(request):
    if request.method == "POST":
        name = request.POST.get('name')
        price = request.POST.get('price')
        description = request.POST.get('description')

        if Product.objects.filter(name=name).exists():
            return redirect('/home_edit')

        new_product = Product.objects.create(
            name=name,
            price=price,
            description=description
        )
        new_product.save()

    return redirect('/home_edit')

@login_required
def edit_product(request):
    if request.method == 'POST':
        new_name = request.POST.get('new_name')
        new_price = request.POST.get('new_price')
        new_description = request.POST.get('new_description')
        pk = request.POST.get('product_id')
        product = get_object_or_404(Product, pk=pk)

        if new_name != "":
            product.name = new_name
        if new_price != "":
            product.price = new_price
        if new_description != "":
            product.description = new_description
        product.save()

    return redirect('/home_edit')


@login_required
def delete_product(request, pk):
    if request.method == 'POST':
        product = get_object_or_404(Product, pk=pk)
        for user in User.objects.all():
            wishlist, _ = Wishlist.objects.get_or_create(user=user)
            wishlist.products.remove(product)
        product.is_active = False
        product.save()
    return redirect('/home_edit')


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('/home')
        else:
            return render(request, 'registration/login.html', context={"failed": True})
    else:
        return render(request, 'registration/login.html')


def info(request):
    if request.user.is_authenticated:
        return redirect('/home')
    return render(request, 'main/info.html')

@login_required
def report(request):
    if not request.user.is_staff:
        return HttpResponseForbidden('У вас нет доступа к этой странице.')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=report.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Отчёт'
    columns = ['Название', 'Стоимость', 'Описание', 'Кол-во приобретений']
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    products = Product.objects.all()
    for product in products:
        purchased_count = Purchase.objects.filter(product=product).count()
        row_num += 1
        row = [
            product.name,
            str(product.price),
            product.description,
            str(purchased_count)
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response

@login_required
def export_benefits_xls(request):
    purchases_by_user = Purchase.objects.filter().select_related('product').order_by('-date')
    workbook = openpyxl.Workbook()

    worksheet = workbook.active
    worksheet.title = 'Purchases'

    worksheet['A1'] = 'Льгота'
    worksheet['B1'] = 'Дата'

    row_num = 2
    for purchase in purchases_by_user:
        worksheet.cell(row=row_num, column=1, value=purchase.product.name)
        worksheet.cell(row=row_num, column=2, value=purchase.date.strftime('%Y-%m-%d %H:%M:%S'))
        row_num += 1

    worksheet.column_dimensions['A'].width = 50
    worksheet.column_dimensions['B'].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=purchases.xlsx'
    workbook.save(response)
    return response
